"""
Approche : CoT (Llama) + Juge GPT-4o-mini (OpenRouter)
—————————————————————————————————————————————————————————
Pipeline :
  1. Llama classifie via Chain-of-Thought (pas d'auto-critique)
  2. GPT-4o-mini joue le rôle de juge : valide ou corrige
  3. Export Excel avec comparaison vs vérité terrain
"""

import os
import re
import json
import time
import random
import traceback
from pathlib import Path
from collections import defaultdict

import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()

# ============================================================
# ⚙️  CONFIGURATION
# ============================================================

# 1. Dossier contenant vos fichiers .md (source des segments)
MARKDOWN_INPUT_DIR = Path(
    r"C:\Users\Jihene\Downloads\Business-Value-Knowledge-Graph\main"
    r"\Extraction_et_Embedding\extract_per_client\Docs"
)

# 2. Dossier de sortie pour les résultats JSON
OUTPUT_DIR = Path(
    r"C:\Users\Jihene\Downloads\Business-Value-Knowledge-Graph\main"
    r"\Extraction_et_Embedding\classification_test\resultats_cot_gpt"
)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# 3. Fichier Excel vérité terrain (multi-LLM)
GROUND_TRUTH_EXCEL = Path(
    r"C:\Users\Jihene\Downloads\Business-Value-Knowledge-Graph\main"
    r"\Extraction_et_Embedding\classification_test\verite_terrain\multi_llm_results.xlsx"
)

# 4. Excel de sortie final
OUTPUT_EXCEL = OUTPUT_DIR / "rapport_cot_gpt_juge.xlsx"

# 5. Clés API
OPENROUTER_API_KEY = os.environ.get("OPENROUTER_API_KEY", "")
if not OPENROUTER_API_KEY:
    raise ValueError("OPENROUTER_API_KEY non définie")

# 6. Modèles
LLAMA_MODEL = "meta-llama/llama-3.3-70b-instruct"   # CoT
JUDGE_MODEL = "openai/gpt-4o-mini"                  # Juge

OPENROUTER_URL = "https://openrouter.ai/api/v1/chat/completions"

# 7. Paramètres
MAX_TOKENS_COT   = 800
MAX_TOKENS_JUDGE = 400
BASE_DELAY       = 1.0
MAX_RETRIES      = 3
BACKOFF_FACTOR   = 2

CATEGORIES = ["ROI", "Notoriété", "Obligation", "Description"]
MIN_SCORE  = 1   # seuil pour choisir une catégorie


# ============================================================
#   PROMPTS
# ============================================================

# ── Prompt Llama (CoT sans auto-critique) ──────────────────
SYSTEM_COT = """Tu es un expert en analyse de documents commerciaux.
Tu réponds UNIQUEMENT en JSON valide, sans markdown, sans explication hors JSON."""

PROMPT_COT = """Section parente : "{section_title}"

Paragraphe :
\"\"\"
{paragraph}
\"\"\"

Classifie ce paragraphe en suivant ce raisonnement pas à pas :

ÉTAPE 1 — Identifie les mots-clés porteurs de sens.
ÉTAPE 2 — Évalue chaque question (oui/non) :
  [ROI-1] Gain financier, réduction de coût, rentabilité mesurable ?
  [ROI-2] Amélioration fonctionnelle (temps, charge, automatisation) comme avantage opérationnel ?
  [ROI-3] Impact sur résultats, ressources ou performance d'une organisation ?
  [NOT-1] Amélioration bien-être, confort, qualité de vie d'un usager ?
  [NOT-2] Label, reconnaissance, attractivité, image positive ?
  [NOT-3] Impact visible ou perçu positivement dans l'environnement ou l'expérience utilisateur ?
  [OBL-1] Nécessité de respecter une norme, loi ou exigence réglementaire ?
  [OBL-2] Mesure de sécurité ou prévention des risques ?
  [OBL-3] Action nécessaire pour éviter danger, sanction ou garantir protection minimale ?
ÉTAPE 3 — Déduis la catégorie principale :
  ROI        → majorité oui sur ROI-1/2/3
  Notoriété  → majorité oui sur NOT-1/2/3
  Obligation → majorité oui sur OBL-1/2/3
  Description → aucune majorité claire

Retourne ce JSON :
{{
  "raisonnement": "<résumé 2-3 phrases>",
  "mots_cles": ["<mot1>", "<mot2>"],
  "reponses": {{
    "roi_1": "<oui|non>", "roi_2": "<oui|non>", "roi_3": "<oui|non>",
    "not_1": "<oui|non>", "not_2": "<oui|non>", "not_3": "<oui|non>",
    "obl_1": "<oui|non>", "obl_2": "<oui|non>", "obl_3": "<oui|non>"
  }},
  "categorie": "<ROI|Notoriété|Obligation|Description>",
  "confiance": <0.0-1.0>,
  "justification": "<une phrase>"
}}"""

# ── Prompt Juge GPT (remplace l'auto-critique) ─────────────
SYSTEM_JUDGE = """Tu es un juge impartial en analyse de valeur commerciale.
Tu réponds UNIQUEMENT en JSON valide, sans markdown, sans explication hors JSON."""

PROMPT_JUDGE = """Tu joues le rôle d'un juge neutre.

Voici la classification initiale d'un paragraphe par un premier modèle (Llama) :

Section parente : "{section_title}"
Paragraphe :
\"\"\"
{paragraph}
\"\"\"

Raisonnement du modèle : {raisonnement}
Catégorie proposée    : {categorie_proposee}
Confiance du modèle   : {confiance}
Réponses aux 9 questions : {reponses_str}

Catégories possibles :
  ROI         → gain financier, réduction de coût, amélioration opérationnelle mesurable
  Notoriété   → bien-être, image, attractivité, expérience utilisateur positive
  Obligation  → norme, loi, sécurité, réglementation, prévention de risque
  Description → contexte neutre, présentation factuelle sans valeur ajoutée identifiable

En tant que juge :
1. Évalue si la catégorie "{categorie_proposee}" est correcte.
2. Si tu n'es pas d'accord, propose la meilleure catégorie parmi les 4.
3. Justifie ta décision en 1-2 phrases.

Retourne UNIQUEMENT ce JSON :
{{
  "accord_avec_modele": <true|false>,
  "categorie_juge": "<ROI|Notoriété|Obligation|Description>",
  "confiance_juge": <0.0-1.0>,
  "justification_juge": "<1-2 phrases>",
  "analyse_critique": "<ce qui a été mal évalué par le modèle>"
}}"""


# ============================================================
# 🧠  FONCTIONS
# ============================================================

def call_llama_cot(segment: dict) -> dict:
    """Appelle Llama pour la classification CoT (sans auto-critique)."""
    prompt = PROMPT_COT.format(
        section_title=segment["section_title"],
        paragraph=segment["paragraph"][:800],
    )
    messages = [
        {"role": "system", "content": SYSTEM_COT},
        {"role": "user",   "content": prompt},
    ]
    return call_openrouter(LLAMA_MODEL, messages, MAX_TOKENS_COT)


def call_gpt_judge(segment: dict, cot_result: dict) -> dict:
    """Appelle GPT-4o-mini pour juger la classification CoT."""
    # Formatage des réponses pour le prompt
    reponses = cot_result.get("reponses", {})
    reponses_str = "\n".join(f"  {k}: {v}" for k, v in reponses.items())

    prompt = PROMPT_JUDGE.format(
        section_title=segment["section_title"],
        paragraph=segment["paragraph"][:600],
        raisonnement=cot_result.get("raisonnement", ""),
        categorie_proposee=cot_result.get("categorie", "Description"),
        confiance=cot_result.get("confiance", 0.5),
        reponses_str=reponses_str,
    )
    messages = [
        {"role": "system", "content": SYSTEM_JUDGE},
        {"role": "user",   "content": prompt},
    ]
    return call_openrouter(JUDGE_MODEL, messages, MAX_TOKENS_JUDGE)


def call_openrouter(model: str, messages: list, max_tokens: int) -> dict:
    """Appel générique à OpenRouter avec gestion des rate limits."""
    headers = {
        "Authorization": f"Bearer {OPENROUTER_API_KEY}",
        "Content-Type": "application/json",
        "HTTP-Referer": "https://github.com/business-value-kg",
        "X-Title": "Business Value Classification",
    }
    payload = {
        "model": model,
        "messages": messages,
        "temperature": 0,
        "max_tokens": max_tokens,
    }

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = requests.post(OPENROUTER_URL, headers=headers,
                                 json=payload, timeout=60)
            resp.raise_for_status()
            raw = resp.json()["choices"][0]["message"]["content"].strip()
            raw = re.sub(r"^```(?:json)?\s*", "", raw)
            raw = re.sub(r"\s*```$", "", raw)
            return json.loads(raw)

        except Exception as e:
            wait = BASE_DELAY * (BACKOFF_FACTOR ** attempt) + random.uniform(0, 0.5)
            if attempt < MAX_RETRIES:
                print(f"      ⚠️  Erreur {model.split('/')[-1]}: {e} — retry dans {wait:.1f}s")
                time.sleep(wait)
            else:
                raise


def parse_segments_from_markdown(md_path: Path) -> list[dict]:
    """
    Parse un fichier .md pour extraire les segments (titre + paragraphe).
    Version simplifiée (à adapter selon votre format).
    """
    with open(md_path, "r", encoding="utf-8") as f:
        text = f.read()

    segments = []
    lines = text.split("\n")
    current_title = "Introduction"
    buffer = []
    para_idx = 0

    for line in lines:
        line = line.strip()
        if line.startswith("#"):
            # Nouveau titre
            if buffer:
                para_text = " ".join(buffer).strip()
                if len(para_text) >= 30:
                    segments.append({
                        "section_title": current_title,
                        "paragraph": para_text,
                        "paragraph_index": para_idx,
                        "source_file": md_path.name,
                        "source_folder": str(md_path.parent),
                    })
                    para_idx += 1
                buffer = []
            current_title = line.lstrip("#").strip()
        elif line:
            buffer.append(line)
        else:
            # Ligne vide
            if buffer:
                para_text = " ".join(buffer).strip()
                if len(para_text) >= 30:
                    segments.append({
                        "section_title": current_title,
                        "paragraph": para_text,
                        "paragraph_index": para_idx,
                        "source_file": md_path.name,
                        "source_folder": str(md_path.parent),
                    })
                    para_idx += 1
                buffer = []

    # Dernier paragraphe
    if buffer:
        para_text = " ".join(buffer).strip()
        if len(para_text) >= 30:
            segments.append({
                "section_title": current_title,
                "paragraph": para_text,
                "paragraph_index": para_idx,
                "source_file": md_path.name,
                "source_folder": str(md_path.parent),
            })

    return segments


def load_ground_truth(excel_path: Path) -> pd.DataFrame:
    """Charge la vérité terrain."""
    df = pd.read_excel(excel_path)
    df.columns = [str(c).strip() for c in df.columns]
    return df


def match_ground_truth(segments: list[dict], gt_df: pd.DataFrame) -> list[dict]:
    """Apparie chaque segment à sa GT."""
    # Construction du lookup
    gt_lookup = {}
    for _, row in gt_df.iterrows():
        try:
            pidx = int(row["paragraph_index"])
        except (ValueError, TypeError):
            continue
        src = str(row["source_file"]).strip()
        gt_lookup[(src, pidx)] = str(row["category"]).strip()

    for seg in segments:
        src = seg.get("source_file", "")
        pidx = seg.get("paragraph_index", -1)
        seg["categorie_gt"] = gt_lookup.get((src, pidx), "N/A")

    return segments


def compute_scores(reponses: dict) -> tuple[int, int, int]:
    roi = sum(1 for k in ["roi_1","roi_2","roi_3"] if reponses.get(k) == "oui")
    not_ = sum(1 for k in ["not_1","not_2","not_3"] if reponses.get(k) == "oui")
    obl = sum(1 for k in ["obl_1","obl_2","obl_3"] if reponses.get(k) == "oui")
    return roi, not_, obl


def determine_label(roi: int, not_: int, obl: int) -> str:
    if roi < MIN_SCORE and not_ < MIN_SCORE and obl < MIN_SCORE:
        return "Description"
    scores = {"ROI": roi, "Obligation": obl, "Notoriété": not_}
    max_score = max(scores.values())
    tied = [c for c, s in scores.items() if s == max_score]
    for priority in ["ROI", "Obligation", "Notoriété"]:
        if priority in tied:
            return priority
    return tied[0] if tied else "Description"


def accuracy(preds: list[str], gts: list[str]) -> float:
    pairs = [(p, g) for p, g in zip(preds, gts) if g != "N/A"]
    if not pairs:
        return 0.0
    return round(sum(p == g for p, g in pairs) / len(pairs) * 100, 1)


# ============================================================
# 📊  EXPORT EXCEL
# ============================================================

def export_excel(segments: list[dict], output_path: Path) -> None:
    """Exporte les résultats en Excel (3 feuilles)."""
    wb = Workbook()

    # Feuille 1 : Détail
    ws1 = wb.active
    ws1.title = "Détail segments"
    headers = [
        "N°", "Fichier", "Section", "Paragraphe (extrait)",
        "Catégorie CoT", "Confiance CoT",
        "Accord juge ?", "Catégorie juge", "Confiance juge",
        "Justification juge", "Analyse critique juge",
        "Vérité terrain",
        "CoT = GT ?", "Juge = GT ?",
    ]
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h).font = Font(bold=True)

    for idx, seg in enumerate(segments, 1):
        row = [
            idx,
            seg.get("source_file", ""),
            seg.get("section_title", "")[:60],
            seg.get("paragraph", "")[:200],
            seg.get("categorie_cot", "?"),
            seg.get("confiance_cot", 0.0),
            "OUI" if seg.get("accord_avec_modele") else "NON",
            seg.get("categorie_juge", "?"),
            seg.get("confiance_juge", 0.0),
            seg.get("justification_juge", ""),
            seg.get("analyse_critique", ""),
            seg.get("categorie_gt", "N/A"),
            "✓" if seg.get("categorie_cot") == seg.get("categorie_gt") else "✗",
            "✓" if seg.get("categorie_juge") == seg.get("categorie_gt") else "✗",
        ]
        for col, val in enumerate(row, 1):
            ws1.cell(row=idx+1, column=col, value=val)

    # Feuille 2 : Statistiques
    ws2 = wb.create_sheet("Stats")
    ws2.cell(row=1, column=1, value="Statistiques de jugement").font = Font(bold=True)

    total = len(segments)
    accords = sum(1 for s in segments if s.get("accord_avec_modele"))
    desaccords = total - accords

    ws2.cell(row=3, column=1, value="Total segments")
    ws2.cell(row=3, column=2, value=total)
    ws2.cell(row=4, column=1, value="Juge en accord")
    ws2.cell(row=4, column=2, value=f"{accords} ({accords/total*100:.1f}%)")
    ws2.cell(row=5, column=1, value="Juge en désaccord")
    ws2.cell(row=5, column=2, value=f"{desaccords} ({desaccords/total*100:.1f}%)")

    # Feuille 3 : Accuracy vs GT
    ws3 = wb.create_sheet("Accuracy_vs_GT")
    segs_gt = [s for s in segments if s.get("categorie_gt") != "N/A"]
    if segs_gt:
        acc_cot = accuracy([s["categorie_cot"] for s in segs_gt],
                           [s["categorie_gt"] for s in segs_gt])
        acc_juge = accuracy([s["categorie_juge"] for s in segs_gt],
                            [s["categorie_gt"] for s in segs_gt])

        ws3.cell(row=1, column=1, value="Modèle")
        ws3.cell(row=1, column=2, value="Accuracy (%)")
        ws3.cell(row=2, column=1, value="CoT (Llama)")
        ws3.cell(row=2, column=2, value=f"{acc_cot}%")
        ws3.cell(row=3, column=1, value="CoT + Juge GPT")
        ws3.cell(row=3, column=2, value=f"{acc_juge}%")
        ws3.cell(row=4, column=1, value="Δ (amélioration)")
        ws3.cell(row=4, column=2, value=f"{acc_juge - acc_cot:+.1f} pp")

    wb.save(output_path)
    print(f"✅ Excel sauvegardé : {output_path}")


# ============================================================
# 🚀  MAIN
# ============================================================

def run():
    print("=" * 80)
    print("  CoT (Llama) + Juge GPT-4o-mini")
    print("  → L'auto-critique est remplacée par le juge externe")
    print("=" * 80)

    # 1. Charger les segments depuis les .md
    print(f"\n📂 Chargement des .md depuis : {MARKDOWN_INPUT_DIR}")
    if not MARKDOWN_INPUT_DIR.exists():
        print(f"  ✗ Dossier introuvable")
        return

    all_segments = []
    for md_file in MARKDOWN_INPUT_DIR.rglob("*.md"):
        print(f"    • {md_file.name}")
        segments = parse_segments_from_markdown(md_file)
        all_segments.extend(segments)

    print(f"  ✓ Total : {len(all_segments)} segments chargés")

    # 2. Charger la vérité terrain (optionnel)
    print(f"\n📊 Chargement de la vérité terrain : {GROUND_TRUTH_EXCEL}")
    if GROUND_TRUTH_EXCEL.exists():
        gt_df = load_ground_truth(GROUND_TRUTH_EXCEL)
        all_segments = match_ground_truth(all_segments, gt_df)
    else:
        print("  ⚠️  Fichier GT introuvable — colonnes GT = N/A")

    # 3. Traitement de chaque segment
    print(f"\n🤖 Traitement des {len(all_segments)} segments...")

    for idx, seg in enumerate(all_segments, 1):
        src = seg.get("source_file", "")
        pid = seg.get("paragraph_index", "?")
        print(f"\n  [{idx:>3}/{len(all_segments)}] {src} | p={pid}")

        # Étape 1 : CoT avec Llama
        print(f"      🔹 CoT (Llama)...", end=" ", flush=True)
        cot_result = call_llama_cot(seg)
        seg["categorie_cot"] = cot_result.get("categorie", "Description")
        seg["confiance_cot"] = cot_result.get("confiance", 0.0)
        seg["raisonnement"] = cot_result.get("raisonnement", "")
        seg["reponses"] = cot_result.get("reponses", {})
        print(f"{seg['categorie_cot']} (conf:{seg['confiance_cot']:.2f})")

        # Calcul des scores depuis les réponses (vérification)
        roi, not_, obl = compute_scores(seg["reponses"])
        label_from_scores = determine_label(roi, not_, obl)
        print(f"          Scores: ROI={roi} NOT={not_} OBL={obl} → label calculé: {label_from_scores}")

        # Étape 2 : Juge GPT (remplace l'auto-critique)
        print(f"      🔹 Juge GPT...", end=" ", flush=True)
        judge_result = call_gpt_judge(seg, cot_result)
        seg["accord_avec_modele"] = judge_result.get("accord_avec_modele")
        seg["categorie_juge"] = judge_result.get("categorie_juge", seg["categorie_cot"])
        seg["confiance_juge"] = judge_result.get("confiance_juge", 0.0)
        seg["justification_juge"] = judge_result.get("justification_juge", "")
        seg["analyse_critique"] = judge_result.get("analyse_critique", "")

        if seg["accord_avec_modele"]:
            print(f"✓ accord (catégorie jugée correcte)")
        else:
            print(f"✗ désaccord → juge propose: {seg['categorie_juge']}")
        print(f"          Justification: {seg['justification_juge'][:100]}...")

        time.sleep(BASE_DELAY)

    # 4. Sauvegarde JSON
    json_path = OUTPUT_DIR / "segments_complets.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(all_segments, f, ensure_ascii=False, indent=2)
    print(f"\n💾 JSON sauvegardé : {json_path}")

    # 5. Export Excel
    export_excel(all_segments, OUTPUT_EXCEL)

    # 6. Résumé
    total = len(all_segments)
    accords = sum(1 for s in all_segments if s.get("accord_avec_modele"))
    segs_gt = [s for s in all_segments if s.get("categorie_gt") != "N/A"]

    print("\n" + "=" * 80)
    print("  RÉSUMÉ FINAL")
    print("=" * 80)
    print(f"  Segments total              : {total}")
    print(f"  Juge en accord avec Llama   : {accords} ({accords/total*100:.1f}%)")
    print(f"  Juge en désaccord           : {total-accords} ({(total-accords)/total*100:.1f}%)")

    if segs_gt:
        acc_cot = accuracy([s["categorie_cot"] for s in segs_gt],
                           [s["categorie_gt"] for s in segs_gt])
        acc_juge = accuracy([s["categorie_juge"] for s in segs_gt],
                            [s["categorie_gt"] for s in segs_gt])
        print(f"\n  Accuracy vs vérité terrain ({len(segs_gt)} segments) :")
        print(f"    CoT (Llama)         : {acc_cot}%")
        print(f"    CoT + Juge GPT-4o   : {acc_juge}%  [Δ = {acc_juge - acc_cot:+.1f} pp]")

    print(f"\n📁 Résultats dans : {OUTPUT_DIR}")


if __name__ == "__main__":
    try:
        run()
    except KeyboardInterrupt:
        print("\n⏹  Interruption")
    except Exception as e:
        print(f"\n✗ Erreur fatale : {e}")
        traceback.print_exc()